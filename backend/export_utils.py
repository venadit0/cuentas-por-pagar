from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
import io

def create_invoices_excel(invoices, estado_filter, empresa_nombre):
    """Crea un archivo Excel con las facturas filtradas"""
    
    # Crear workbook
    wb = Workbook()
    ws = wb.active
    
    # Configurar título de la hoja
    titulo = f"Facturas {estado_filter.title()}"
    ws.title = titulo[:31]  # Excel limita a 31 caracteres
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Información de la empresa (primeras filas)
    ws.merge_cells('A1:F1')
    ws['A1'] = f"REPORTE DE FACTURAS - {empresa_nombre}"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal="center")
    
    ws.merge_cells('A2:F2')
    ws['A2'] = f"Estado: {estado_filter.title()}"
    ws['A2'].font = Font(bold=True, size=12)
    ws['A2'].alignment = Alignment(horizontal="center")
    
    ws.merge_cells('A3:F3')
    ws['A3'] = f"Fecha de generación: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws['A3'].alignment = Alignment(horizontal="center")
    
    # Fila vacía
    ws['A4'] = ""
    
    # Encabezados (fila 5)
    headers = [
        "Número de Factura",
        "Número de Contrato",
        "Proveedor", 
        "Fecha de Factura",
        "Monto",
        "Estado de Pago",
        "Archivo PDF"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Datos de facturas (desde fila 6)
    for row, invoice in enumerate(invoices, 6):
        ws.cell(row=row, column=1, value=invoice.get('numero_factura', ''))
        ws.cell(row=row, column=2, value=invoice.get('numero_contrato', '') or 'Sin asignar')
        ws.cell(row=row, column=3, value=invoice.get('nombre_proveedor', ''))
        
        # Formatear fecha
        fecha_str = invoice.get('fecha_factura', '')
        if fecha_str:
            try:
                fecha_obj = datetime.fromisoformat(fecha_str.replace('Z', '+00:00')) if 'T' in fecha_str else datetime.strptime(fecha_str, '%Y-%m-%d')
                ws.cell(row=row, column=4, value=fecha_obj.strftime('%d/%m/%Y'))
            except:
                ws.cell(row=row, column=4, value=fecha_str)
        
        # Formatear monto
        monto = invoice.get('monto', 0)
        ws.cell(row=row, column=5, value=f"${monto:,.2f}")
        
        ws.cell(row=row, column=6, value=invoice.get('estado_pago', '').title())
        ws.cell(row=row, column=7, value=invoice.get('archivo_original', invoice.get('archivo_pdf', 'N/A')))
    
    # Ajustar ancho de columnas
    column_widths = [18, 20, 25, 15, 15, 15, 20]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + col)].width = width
    
    # Agregar totales si hay facturas
    if invoices:
        total_row = len(invoices) + 7
        ws.merge_cells(f'A{total_row}:C{total_row}')
        ws[f'A{total_row}'] = f"TOTAL ({len(invoices)} facturas):"
        ws[f'A{total_row}'].font = Font(bold=True)
        ws[f'A{total_row}'].alignment = Alignment(horizontal="right")
        
        total_monto = sum(invoice.get('monto', 0) for invoice in invoices)
        ws[f'D{total_row}'] = f"${total_monto:,.2f}"
        ws[f'D{total_row}'].font = Font(bold=True)
    
    # Guardar en memoria
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    return excel_buffer

def create_summary_excel(resumen_data, empresa_nombre):
    """Crea un archivo Excel con resumen general"""
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen General"
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    title_font = Font(bold=True, size=14)
    
    # Título
    ws.merge_cells('A1:E1')
    ws['A1'] = f"RESUMEN FINANCIERO - {empresa_nombre}"
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal="center")
    
    ws.merge_cells('A2:E2')
    ws['A2'] = f"Fecha: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws['A2'].alignment = Alignment(horizontal="center")
    
    # Resumen general (fila 4)
    ws['A4'] = "RESUMEN GENERAL"
    ws['A4'].font = Font(bold=True, size=12)
    
    ws['A5'] = "Deuda Total Pendiente:"
    ws['B5'] = f"${resumen_data.get('total_deuda_global', 0):,.2f}"
    ws['B5'].font = Font(bold=True, color="FF0000")
    
    ws['A6'] = "Total de Facturas:"
    ws['B6'] = resumen_data.get('total_facturas', 0)
    
    ws['A7'] = "Facturas Pendientes:"
    ws['B7'] = resumen_data.get('facturas_pendientes', 0)
    
    ws['A8'] = "Facturas Pagadas:"
    ws['B8'] = resumen_data.get('facturas_pagadas', 0)
    
    # Resumen por proveedor (fila 10)
    ws['A10'] = "RESUMEN POR PROVEEDOR"
    ws['A10'].font = Font(bold=True, size=12)
    
    # Encabezados proveedores
    headers = ["Proveedor", "Deuda Pendiente", "Fact. Pendientes", "Fact. Pagadas"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=11, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    
    # Datos proveedores
    proveedores = resumen_data.get('proveedores', [])
    for row, proveedor in enumerate(proveedores, 12):
        ws.cell(row=row, column=1, value=proveedor.get('proveedor', ''))
        ws.cell(row=row, column=2, value=f"${proveedor.get('total_deuda', 0):,.2f}")
        ws.cell(row=row, column=3, value=proveedor.get('facturas_pendientes', 0))
        ws.cell(row=row, column=4, value=proveedor.get('facturas_pagadas', 0))
    
    # Ajustar anchos
    column_widths = [25, 18, 18, 18]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + col)].width = width
    
    # Guardar en memoria
    excel_buffer = io.BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)
    
    return excel_buffer